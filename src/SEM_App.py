# Name:			SEM | Soccer Excel Merger
# Version:		0.1.2
# Description:	Import multiplte soccer data excel files and add to larger (season) SEM dataset excel sheet.

import tkinter as tk
from tkinter import filedialog, messagebox, PhotoImage
import os
from openpyxl import Workbook, load_workbook

# Set
heads = ['club',
		 'opponent',
		 'win',
		 'draw',
		 'clean sheet',
		 'goal',
		 'ball possession',
		 'goal attempt',
		 'corner',
		 'goalkeeper save',
		 'foal',
		 'red card',
		 'yellow card',
		 'penalty save',
		 'total points',
		 'home',
		 'zworth',
		 'nummer',
		 'round'
		 ]
guiTitle = "Soccer Excel Merger 0.1.2"
guiHeight = 500
guiWidth = 500
guiSize = str(guiWidth) + "x" + str(guiHeight)
guiColor1 = "#40514E"
guiColor2 = "#11999E"
guiColor3 = "#E4F9F5"
guiColor4 = "#C9CBC9"
filePathImports = []
datasetPathImport = ''
programLocation = os.path.abspath(__file__)

# Functions
def select_import_files():

	filespath = filedialog.askopenfilenames(initialdir=programLocation, title="Select file", filetypes=(("Microsoft Excel Worksheet", "*.xlsx"), ("All Files", "*,*")))

	if filespath:
		global filePathImports
		filePathImports = list(filespath)

		# Check for duplicates
		# ...

	importFilesListbox.delete(0, tk.END)

	for file in filePathImports:
		importFilesListbox.insert(tk.END, os.path.basename(file))


def clear_import_files():

	global filePathImports
	filePathImports = []

	importFilesListbox.delete(0, tk.END)


def select_dataset_file():
	
	filepath = filedialog.askopenfilename(initialdir=programLocation, title="Select file", filetypes=(("Microsoft Excel Worksheet", "*.xlsx"), ("All Files", "*,*")))

	if filepath:
		global datasetPathImport
		datasetPathImport = filepath

	datasetListbox.delete(0, tk.END)

	datasetListbox.insert(tk.END, os.path.basename(filepath))


def clear_dataset_file():
	
	global datasetPathImport
	datasetPathImport = ''

	datasetListbox.delete(0, tk.END)


def extract_club_names(namesString, clubsList):
	tempTeam = []

	for name in clubsList:

		nameIndex = namesString.find(name)
		if nameIndex > -1:

			if len(tempTeam) > 0:

				# check name against name pos
				if tempTeam[1] < nameIndex:
					return [tempTeam[0], name]
				else:
					return [name, tempTeam[0]]
			else:
				tempTeam = [name, nameIndex]

def string_clean_int(string):
	return int(str(string).split(" ")[0])


def assign_results(wb, sheetName, matchNames, home, playRound):
	
	tempResults = []
	for cell in wb[sheetName]['B']:
		if cell.value == None:
			tempResults.append(0)
		else:
			tempResults.append(cell.value)

	rowPlus = 0 if home else 14

	tempClub = {}
	tempClubName = matchNames[1 - home]

	tempClub['club'] = tempClubName
	tempClub['opponent'] = matchNames[home]

	tempClub['win'] = string_clean_int(tempResults[2 + rowPlus])
	tempClub['draw'] = string_clean_int(tempResults[3 + rowPlus])
	tempClub['clean sheet'] = string_clean_int(tempResults[4 + rowPlus])
	tempClub['goal'] = string_clean_int(tempResults[5 + rowPlus])
	tempClub['ball possession'] = string_clean_int(tempResults[6 + rowPlus]) # sdfsdf
	tempClub['goal attempt'] = string_clean_int(tempResults[7 + rowPlus])
	tempClub['corner'] = string_clean_int(tempResults[8 + rowPlus])
	tempClub['goalkeeper save'] = string_clean_int(tempResults[9 + rowPlus])
	tempClub['foal'] = string_clean_int(tempResults[10 + rowPlus])
	tempClub['red card'] = string_clean_int(tempResults[11 + rowPlus])
	tempClub['yellow card'] = string_clean_int(tempResults[12 + rowPlus])
	tempClub['penalty save'] = string_clean_int(tempResults[13 + rowPlus])
	tempClub['total points'] = string_clean_int(tempResults[14 + rowPlus])

	tempClub['home'] = home
	tempClub['round'] = playRound

	for i in range(1, 19):
			if wb['values'].cell(row=i, column=3).value == tempClubName:
				tempClub['zworth'] = float(str(wb['values'].cell(row=i, column=5).value).split(" ")[0])
				tempClub['nummer'] = int(wb['values'].cell(row=i, column=1).value)

	# Make list (temp solution)
	listClub = []
	for head in heads:
		listClub.append(tempClub[head])
	tempClub['list'] = listClub
	
	return tempClub


def extract_import_results(dataset = 0):

	tempRowsList = []

	clubNames = []
	for cell in load_workbook(filePathImports[0])['values']['C']:
		clubNames.append(cell.value)

	numClubs = len(clubNames)

	if dataset != 0 and 'results' in dataset.sheetnames:
		playRound = int((dataset['results'].max_row - 1) / numClubs)

	elif dataset != 0 and 'points' in dataset.sheetnames:
		playRound = int((dataset['points'].max_row - 1) / numClubs)

	else:
		playRound = 0


	for filePath in filePathImports:
		wbImport = load_workbook(filePath)

		playRound += 1

		for sheet in wbImport.sheetnames:
			if sheet != 'values':

				# Check for good paste sheets and values, else error message

				matchNames = extract_club_names(wbImport[sheet]['A1'].value, clubNames)

				tempRowsList.append(assign_results(wbImport, sheet, matchNames, 1, playRound))
				tempRowsList.append(assign_results(wbImport, sheet, matchNames, 0, playRound))

	return tempRowsList


def convert_to_points(results):

	for row in results:

		row['points'] = [row['club'], row['opponent']]
		row['points'].append(row['win'] * 100)
		row['points'].append(row['draw'] * 50)
		row['points'].append(row['clean sheet'] * 30)
		row['points'].append(row['goal'] * 20)
		row['points'].append(0 if row['ball possession'] < 60 else 20)
		row['points'].append(row['goal attempt'] * 2)
		row['points'].append(row['corner'] * 3)
		row['points'].append(row['goalkeeper save'] * 4)
		row['points'].append(row['foal'] * -1)
		row['points'].append(row['red card'] * -10)
		row['points'].append(row['yellow card'] * -5)
		row['points'].append(row['penalty save'] * 20)
		row['points'] += [row['total points'], row['home'], row['zworth'], row['nummer'], row['round']]

	return results


def merge_save_dataset(wbDataset, saveFilePath):

	# Check wb tabs
	if checkResults.get():
		if 'results' not in wbDataset.sheetnames:
			tk.messagebox.showerror('Alert', 'The dataset has no results tab')
			return

	if checkPoints.get():
		if 'points' not in wbDataset.sheetnames:
			tk.messagebox.showerror('Alert', 'The dataset has no points tab')
			return
	
	# Import results
	resultsList = extract_import_results(wbDataset)

	# Merge
	if checkResults.get():
		for row in resultsList:
			wbDataset['results'].append(row['list'])

	if checkPoints.get():
		resultsList = convert_to_points(resultsList)
		for row in resultsList:
			wbDataset['points'].append(row['points'])

	# Save
	wbDataset.save(saveFilePath)

	return True


def add_to_dataset():
	
	# Check
	if not filePathImports:
		tk.messagebox.showerror('Alert', 'No import file(s) selected.')
		return

	if not datasetPathImport:
		tk.messagebox.showerror('Alert', 'No dataset file selected.')
		return

	if not checkResults.get() and not checkPoints.get():
		tk.messagebox.showerror('Alert', 'No results or points checked')
		return
	
	excWb = load_workbook(datasetPathImport)

	if merge_save_dataset(excWb, datasetPathImport):
		tk.messagebox.showinfo('Saved', 'Succesfully added to existing dataset!')


def create_new_dataset():

	tempwb = Workbook()

	if checkResults.get() and checkPoints.get():
		tempwb.active.title = "results"
		tempwb["results"].append(heads)
		tempwb.create_sheet("points")
		tempwb["points"].append(heads)

	elif checkResults.get() and not checkPoints.get():
		tempwb.active.title = "results"
		tempwb["results"].append(heads)

	elif checkPoints.get() and not checkResults.get():
		tempwb.active.title = "points"
		tempwb["points"].append(heads)

	return tempwb

def save_new_dataset():

	# Check (Not needed anymore because of check_points, check_results)
	if not checkResults.get() and not checkPoints.get():
		tk.messagebox.showerror('Alert', 'No results or points checked')
		return

	# Load
	filePath = filedialog.asksaveasfilename(initialdir=programLocation, title="Select file", filetypes=(("Microsoft Excel Worksheet", "*.xlsx"), ("All Files", "*,*")))

	if not filePath:
		return

	# Check extension
	if os.path.splitext(filePath)[-1].lower() != '.xlsx':
		filePath += '.xlsx'

	# if 1 1 - Copy/new + import
	if filePathImports and datasetPathImport:

		excWb = load_workbook(datasetPathImport)

		if merge_save_dataset(excWb, filePath):
			tk.messagebox.showinfo('Saved', 'Succesfully added and saved a copy of the dataset with imports!')

	# if 0 0 - New
	elif not filePathImports and not datasetPathImport:

		newWb = create_new_dataset()

		newWb.save(filePath)
		tk.messagebox.showinfo('Saved', 'Succesfully created a new empty dataset!')

	# if 1 0 - New + import
	elif filePathImports and not datasetPathImport:

		newWb = create_new_dataset()

		if merge_save_dataset(newWb, filePath):
			tk.messagebox.showinfo('Saved', 'Succesfully created a new dataset with imports!')

	elif not filePathImports and datasetPathImport:

		tk.messagebox.showerror('Alert', 'No import file(s) selected.')


def check_points():
	if not checkResults.get() and not checkPoints.get():
		exportResultsCheck.select()

def check_results():
	if not checkResults.get() and not checkPoints.get():
		exportPointsCheck.select()


# GUI
root = tk.Tk()
root.title(guiTitle)
# root.geometry("guiSize")
root.configure(bg=guiColor1)
icon = PhotoImage(file = 'icon.png')
root.iconphoto(False, icon)

root.grid_rowconfigure(0, minsize=20)
root.grid_columnconfigure(0, minsize=20)

selectImportFiles = tk.Button(root, text="Open Import File(s)", width=18, fg=guiColor3, bg=guiColor2, command=select_import_files)
selectImportFiles.grid(row=1, column=1, sticky=tk.N)
importFilesListbox = tk.Listbox(root, height=10, width=40, selectbackground=guiColor2, selectmode=tk.EXTENDED)
importFilesListbox.grid(row=1, column=3, rowspan=10, columnspan=2)
clearImportFiles = tk.Button(root, text="Clear File(s)", width=18, fg=guiColor3, bg=guiColor2, command=clear_import_files)
clearImportFiles.grid(row=2, column=1)

root.grid_rowconfigure(11, minsize=20)

selectDataset = tk.Button(root, text="Open Dataset", width=18, fg=guiColor3, bg=guiColor2, command=select_dataset_file)
selectDataset.grid(row=12, column=1)
datasetListbox = tk.Listbox(root, height=3, width=40, selectbackground=guiColor2, selectmode=tk.SINGLE)
datasetListbox.grid(row=12, column=3, rowspan=2, columnspan=2)
clearDataset = tk.Button(root, text="Clear Dataset", width=18, fg=guiColor3, bg=guiColor2, command=clear_dataset_file)
clearDataset.grid(row=13, column=1)

root.grid_rowconfigure(14, minsize=10)

exportLabel = tk.Label(root, text="Export:", fg=guiColor3, bg=guiColor1)
exportLabel.grid(row=15, column=1)
checkResults = tk.BooleanVar() 
exportResultsCheck = tk.Checkbutton(root, text="Results", fg=guiColor3, bg=guiColor1, activebackground=guiColor1, activeforeground=guiColor3, selectcolor=guiColor1, variable=checkResults, command=check_results)
exportResultsCheck.grid(row=15, column=3)
exportResultsCheck.select()
checkPoints = tk.BooleanVar() 
exportPointsCheck = tk.Checkbutton(root, text="Points", fg=guiColor3, bg=guiColor1, activebackground=guiColor1, activeforeground=guiColor3, selectcolor=guiColor1, variable=checkPoints, command=check_points)
exportPointsCheck.grid(row=15, column=4)
exportPointsCheck.select()

root.grid_rowconfigure(16, minsize=10)

saveDataset = tk.Button(root, text="Save new/copy", width=16, fg=guiColor3, bg=guiColor2, command=save_new_dataset)
saveDataset.grid(row=17, column=3)
addDataset = tk.Button(root, text="Add", width=16, fg=guiColor3, bg=guiColor2, command=add_to_dataset)
addDataset.grid(row=17, column=4)

root.grid_rowconfigure(18, minsize=20)
root.grid_columnconfigure(2, minsize=20)
root.grid_columnconfigure(5, minsize=20)

# frame = tk.Frame(root, bg="gray")
# frame.place(relwidth=0.8, relheight=0.8, relx=0.1, rely=0.1)

# Run GUI
root.mainloop()